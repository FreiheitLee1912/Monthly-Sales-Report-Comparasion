from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.reports import (
    generate_excel_summary_report,
    generate_monthly_compare_report,
    generate_student_gantt_report,
)


def test_excel_summary_report_creates_all_formats(tmp_path):
    source = tmp_path / "sales.xlsx"
    pd.DataFrame({"customer": ["A", "B"], "amount": [100, 250]}).to_excel(source, index=False)

    result = generate_excel_summary_report(source, tmp_path / "out")

    assert result.title == "Excel Summary Report"
    assert result.xlsx_path.exists()
    assert result.html_path.exists()
    assert result.pdf_path.exists()
    assert "sales.xlsx" in result.html_path.read_text(encoding="utf-8")


def test_monthly_compare_report_counts_added_removed_and_changed_rows(tmp_path):
    old_file = tmp_path / "old.xlsx"
    new_file = tmp_path / "new.xlsx"
    pd.DataFrame(
        {
            "A": ["old common", "removed row", "ignored no key"],
            "B": ["x", "y", "z"],
            "C": ["c", "c", "c"],
            "KEY": ["K1", "K2", ""],
            "Amount": [100, 200, 999],
            "BPCS_CUSTOMER_CODE": ["OLD-IGNORED", "R", "I"],
        }
    ).to_excel(old_file, index=False)
    pd.DataFrame(
        {
            "A": ["new common", "added row"],
            "B": ["x", "n"],
            "C": ["c", "c"],
            "KEY": ["K1", "K3"],
            "Amount": [150, 300],
            "BPCS_CUSTOMER_CODE": ["NEW-IGNORED", "A"],
            "NewField": ["new column", "new column"],
        }
    ).to_excel(new_file, index=False)

    result = generate_monthly_compare_report(old_file, new_file, tmp_path / "compare")

    html = result.html_path.read_text(encoding="utf-8")
    assert result.xlsx_path.exists()
    assert result.added == 1
    assert result.removed == 1
    assert result.changed_rows == 1
    assert result.changed_fields == 2
    assert "Additions" in html
    assert "Deletions" in html
    assert "Modified Rows" in html

    workbook = load_workbook(result.xlsx_path)
    assert workbook.sheetnames[:5] == ["Dashboard", "Summary", "Comparison", "new data", "old data"]
    assert (result.xlsx_path.parent / "report.xlsx").exists()

    dashboard = workbook["Dashboard"]
    summary = workbook["Summary"]
    assert dashboard["A1"].value == "Item"
    assert dashboard["A2"].value == "Key Metrics"
    assert dashboard["B3"].value == 2
    assert summary["A1"].value == "Type"
    assert summary["A2"].value == "Added"
    assert summary["A3"].value == "Removed"
    assert summary["A4"].value == "Changed"

    new_data = workbook["new data"]
    old_data = workbook["old data"]
    comparison = workbook["Comparison"]
    assert new_data["A3"].fill.fgColor.rgb == "00FFFF00"
    assert old_data["A3"].fill.fgColor.rgb == "00A6A6A6"
    assert new_data["F2"].fill.fgColor.rgb == "00FFA500"
    assert comparison["E2"].fill.fgColor.rgb == "00FFA500"
    assert new_data["G1"].fill.fgColor.rgb == "00D9D9D9"


def test_student_gantt_report_accepts_chinese_columns(tmp_path):
    source = tmp_path / "students.xlsx"
    pd.DataFrame(
        {
            "學生": ["Lee", "Lee", "Chen"],
            "任務": ["Design", "Build", "Review"],
            "開始日": ["2026-06-01", "2026-06-03", "2026-06-02"],
            "結束日": ["2026-06-02", "2026-06-07", "2026-06-04"],
            "進度": [100, 60, 30],
        }
    ).to_excel(source, index=False)

    result = generate_student_gantt_report(source, tmp_path / "gantt")

    html = result.html_path.read_text(encoding="utf-8")
    assert result.xlsx_path.exists()
    assert result.pdf_path.exists()
    assert "Student Gantt Report" in html
    assert "Lee" in html
