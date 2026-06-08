from pathlib import Path
from zipfile import ZipFile

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.workbook_sanitize import compact_workbook_for_legacy_compare, sanitize_workbook_for_openpyxl


def test_sanitize_workbook_removes_pivot_cache_parts_and_relationships(tmp_path):
    source = tmp_path / "source.xlsx"
    pd.DataFrame({"A": ["x"], "B": ["y"], "C": ["z"], "KEY": ["K1"]}).to_excel(source, index=False)

    with ZipFile(source, "a") as z:
        z.writestr("xl/pivotCache/pivotCacheDefinition1.xml", "<pivotCacheDefinition />")
        z.writestr("xl/pivotTables/pivotTable1.xml", "<pivotTableDefinition />")
        z.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" Target="https://example.com"/>
  <Relationship Id="rIdPivot" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" Target="pivotCache/pivotCacheDefinition1.xml"/>
</Relationships>""",
        )

    sanitized = sanitize_workbook_for_openpyxl(source, tmp_path / "sanitized")

    with ZipFile(sanitized) as z:
        names = z.namelist()
        assert "xl/pivotCache/pivotCacheDefinition1.xml" not in names
        assert "xl/pivotTables/pivotTable1.xml" not in names
        rels = z.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8")
        assert "pivotCacheDefinition" not in rels
        assert "https://example.com" in rels


def test_compact_workbook_keeps_only_legacy_compare_sheets(tmp_path):
    source = tmp_path / "large_source.xlsx"
    workbook = Workbook()
    export = workbook.active
    export.title = "Export"
    export.append(["A", "B", "C", "KEY", "Amount"])
    export.append(["row", "x", "c", "K1", 100])
    noise = workbook.create_sheet("Large irrelevant sheet")
    noise.append(["this sheet should not be copied"])
    history = workbook.create_sheet("LINE_NO history")
    history.append(["Add", "Remove"])
    history.append(["K9", "K8"])
    export2 = workbook.create_sheet("Export (2)")
    export2.append(["A", "B", "C", "KEY", "Amount"])
    export2.append(["row2", "x", "c", "K2", 200])
    workbook.save(source)

    compacted = compact_workbook_for_legacy_compare(source, tmp_path / "compact")

    compacted_workbook = load_workbook(compacted, data_only=True)
    assert compacted_workbook.sheetnames == ["Export", "LINE_NO history", "Export (2)"]
    assert compacted_workbook["Export"]["D2"].value == "K1"
    assert compacted_workbook["Export (2)"]["D2"].value == "K2"
    assert "Large irrelevant sheet" not in compacted_workbook.sheetnames
