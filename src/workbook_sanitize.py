from __future__ import annotations

import re
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook


PIVOT_PART_RE = re.compile(r"(^|/)(pivotCache|pivotTables?)/|pivotCache|pivotTable", re.IGNORECASE)
PIVOT_REL_RE = re.compile(r"pivot(CacheDefinition|Table)", re.IGNORECASE)
LINE_NO_RE = re.compile(r"^LINE_NO", re.IGNORECASE)


def sanitize_workbook_for_openpyxl(source: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    target = output_dir / source.name
    with ZipFile(source, "r") as zin, ZipFile(target, "w", compression=ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            name = item.filename
            if PIVOT_PART_RE.search(name):
                continue
            data = zin.read(item)
            if name.endswith(".rels"):
                data = _remove_pivot_relationships(data)
            elif name.endswith(".xml"):
                data = _remove_pivot_xml_nodes(data)
            zout.writestr(item, data)
    return target


def compact_workbook_for_legacy_compare(source: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    target = output_dir / source.name
    source_workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        selected_sheet_names = _legacy_compare_sheet_names(source_workbook.sheetnames)
        target_workbook = Workbook(write_only=True)
        for sheet_name in selected_sheet_names:
            source_sheet = source_workbook[sheet_name]
            target_sheet = target_workbook.create_sheet(title=sheet_name)
            _copy_sheet_values(source_sheet, target_sheet)
        target_workbook.save(target)
    finally:
        source_workbook.close()
    return target


def _remove_pivot_relationships(data: bytes) -> bytes:
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return data
    removed = False
    for child in list(root):
        rel_type = child.attrib.get("Type", "")
        target = child.attrib.get("Target", "")
        if PIVOT_REL_RE.search(rel_type) or PIVOT_REL_RE.search(target):
            root.remove(child)
            removed = True
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True) if removed else data


def _remove_pivot_xml_nodes(data: bytes) -> bytes:
    try:
        root = ElementTree.fromstring(data)
    except ElementTree.ParseError:
        return data
    removed = _remove_matching_nodes(root)
    return ElementTree.tostring(root, encoding="utf-8", xml_declaration=True) if removed else data


def _remove_matching_nodes(parent) -> bool:
    removed = False
    for child in list(parent):
        tag_name = child.tag.rsplit("}", 1)[-1]
        if tag_name.startswith("pivot"):
            parent.remove(child)
            removed = True
        else:
            removed = _remove_matching_nodes(child) or removed
    return removed


def _legacy_compare_sheet_names(sheet_names: list[str]) -> list[str]:
    selected = []
    preferred_first = "Export" if "Export" in sheet_names else sheet_names[0]
    for sheet_name in [preferred_first, *sheet_names]:
        normalized = sheet_name.replace(" ", "").strip().upper()
        if sheet_name == preferred_first or sheet_name == "Export" or sheet_name == "Export (2)" or LINE_NO_RE.match(normalized):
            if sheet_name not in selected:
                selected.append(sheet_name)
    return selected


def _copy_sheet_values(source_sheet, target_sheet) -> None:
    for row in source_sheet.iter_rows(values_only=True):
        values = list(row)
        while values and values[-1] is None:
            values.pop()
        target_sheet.append(values)
