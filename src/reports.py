from __future__ import annotations

from dataclasses import dataclass
from html import escape
from pathlib import Path
from shutil import copyfile
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib import colors
from src.legacy_compare import compare_sheets as legacy_compare_sheets
from src.workbook_sanitize import sanitize_workbook_for_openpyxl


@dataclass(frozen=True)
class ReportResult:
    title: str
    job_id: str
    xlsx_path: Path
    html_path: Path
    pdf_path: Path
    added: int = 0
    removed: int = 0
    duplication_add: int = 0
    duplication_remove: int = 0
    duplication: int = 0
    changed_rows: int = 0
    changed_fields: int = 0


EXCLUDED_COMPARE_COLUMNS = {
    "BPCS_CUSTOMER_CODE",
    "BPCS_SHIPTO_CODE",
    "CUSTOMER_PN",
    "BUSINESS_CATEGORY_NAME",
    "TRANSACTION_CURRENCY",
    "Ex_Rate_JPY",
    "GENERAL_CODE_1",
    "GENERAL_CODE_2",
    "GENERAL_CODE_3",
    "GENERAL_CODE_4",
    "GENERAL_CODE_5",
}

KEY_COLUMN_INDEX = 4
HEADER_ROW = 1
DATA_START_ROW = 2

FILL_ADDED = PatternFill("solid", fgColor="FFF2CC")
FILL_REMOVED = PatternFill("solid", fgColor="D9E2F3")
FILL_CHANGED = PatternFill("solid", fgColor="FCE4D6")
FILL_NEW_COLUMN = PatternFill("solid", fgColor="E2F0D9")
FILL_EXCLUDED = PatternFill("solid", fgColor="E7E6E6")
FILL_HEADER = PatternFill("solid", fgColor="172B4D")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generate_excel_summary_report(source: Path, output_root: Path) -> ReportResult:
    job_dir = _make_job_dir(output_root)
    sheets = pd.read_excel(source, sheet_name=None)
    rows = []
    for sheet_name, frame in sheets.items():
        rows.append(
            {
                "Sheet": sheet_name,
                "Rows": len(frame),
                "Columns": len(frame.columns),
                "Empty cells": int(frame.isna().sum().sum()),
                "Numeric columns": len(frame.select_dtypes(include="number").columns),
            }
        )
    summary = pd.DataFrame(rows)
    title = "Excel Summary Report"
    return _write_report(title, job_dir, [("Summary", summary)], _summary_html(title, source.name, summary))


def generate_monthly_compare_report(old_file: Path, new_file: Path, output_root: Path) -> ReportResult:
    job_dir = _make_job_dir(output_root)
    sanitized_dir = job_dir / "sanitized_uploads"
    sanitized_old = sanitize_workbook_for_openpyxl(old_file, sanitized_dir / "old")
    sanitized_new = sanitize_workbook_for_openpyxl(new_file, sanitized_dir / "new")
    legacy_result = legacy_compare_sheets(sanitized_old, sanitized_new, job_dir)
    xlsx_path = Path(legacy_result["output_path"])
    download_xlsx_path = job_dir / "report.xlsx"
    if xlsx_path != download_xlsx_path:
        copyfile(xlsx_path, download_xlsx_path)
    html_path = job_dir / "report.html"
    pdf_path = job_dir / "report.pdf"
    title = "Monthly Compare Report"
    summary = pd.DataFrame(
        [
            {"Metric": "Additions", "Count": legacy_result["added"]},
            {"Metric": "Deletions", "Count": legacy_result["removed"]},
            {"Metric": "Duplication in Add", "Count": legacy_result["duplication_add"]},
            {"Metric": "Duplication in Removed", "Count": legacy_result["duplication_remove"]},
            {"Metric": "Duplication Total", "Count": legacy_result["duplication"]},
            {"Metric": "Modified Rows", "Count": legacy_result["changed_rows"]},
            {"Metric": "Modified Fields", "Count": legacy_result["changed_fields"]},
        ]
    )
    html = _summary_html(title, f"{old_file.name} -> {new_file.name}", summary)
    html_path.write_text(html, encoding="utf-8")
    _write_pdf(pdf_path, title, [("Summary", summary)])
    return ReportResult(
        title=title,
        job_id=job_dir.name,
        xlsx_path=xlsx_path,
        html_path=html_path,
        pdf_path=pdf_path,
        added=legacy_result["added"],
        removed=legacy_result["removed"],
        duplication_add=legacy_result["duplication_add"],
        duplication_remove=legacy_result["duplication_remove"],
        duplication=legacy_result["duplication"],
        changed_rows=legacy_result["changed_rows"],
        changed_fields=legacy_result["changed_fields"],
    )


def _generate_monthly_compare_report_reimplementation(old_file: Path, new_file: Path, output_root: Path) -> ReportResult:
    job_dir = _make_job_dir(output_root)
    old_wb = load_workbook(old_file, data_only=False)
    new_wb = load_workbook(new_file, data_only=False)
    new_ws = new_wb.worksheets[0]
    old_ws = _matching_old_sheet(old_wb, new_ws.title)

    old_headers = _headers(old_ws)
    new_headers = _headers(new_ws)
    old_map, old_duplicates = _key_map(old_ws)
    new_map, new_duplicates = _key_map(new_ws)
    old_keys = set(old_map)
    new_keys = set(new_map)
    added_keys = new_keys - old_keys
    removed_keys = old_keys - new_keys
    common_keys = old_keys & new_keys
    comparable_columns = _comparable_columns(old_headers, new_headers)

    detail_rows = []
    changed_keys = set()
    for key in sorted(added_keys):
        detail_rows.append({"Type": "Added", "Line No": new_map[key], "Key": key, "Field": "", "Old Value": "", "New Value": ""})
    for key in sorted(removed_keys):
        detail_rows.append({"Type": "Removed", "Line No": old_map[key], "Key": key, "Field": "", "Old Value": "", "New Value": ""})
    for key in sorted(common_keys):
        old_row = old_map[key]
        new_row = new_map[key]
        for col_idx, field_name in comparable_columns:
            old_value = old_ws.cell(row=old_row, column=col_idx).value
            new_value = new_ws.cell(row=new_row, column=col_idx).value
            if _normalized_cell(old_value) != _normalized_cell(new_value):
                changed_keys.add(key)
                detail_rows.append(
                    {
                        "Type": "Changed",
                        "Line No": new_row,
                        "Key": key,
                        "Field": field_name,
                        "Old Value": "" if old_value is None else old_value,
                        "New Value": "" if new_value is None else new_value,
                    }
                )

    changed_fields = sum(1 for row in detail_rows if row["Type"] == "Changed")
    summary = pd.DataFrame(
        [
            {"Metric": "Previous Period Total", "Count": len(old_map)},
            {"Metric": "Additions", "Count": len(added_keys)},
            {"Metric": "Deletions", "Count": len(removed_keys)},
            {"Metric": "Current Period Total", "Count": len(new_map)},
            {"Metric": "Modified Rows", "Count": len(changed_keys)},
            {"Metric": "Modified Fields", "Count": changed_fields},
        ]
    )
    details = pd.DataFrame(detail_rows, columns=["Type", "Line No", "Key", "Field", "Old Value", "New Value"])
    title = "Monthly Compare Report"
    html = _summary_html(title, f"{old_file.name} -> {new_file.name}", summary, details)
    xlsx_path = job_dir / "report.xlsx"
    html_path = job_dir / "report.html"
    pdf_path = job_dir / "report.pdf"
    _write_compare_workbook(
        xlsx_path=xlsx_path,
        old_ws=old_ws,
        new_ws=new_ws,
        old_headers=old_headers,
        new_headers=new_headers,
        old_map=old_map,
        new_map=new_map,
        added_keys=added_keys,
        removed_keys=removed_keys,
        changed_detail_rows=detail_rows,
        summary=summary,
        details=details,
    )
    html_path.write_text(html, encoding="utf-8")
    _write_pdf(pdf_path, title, [("Summary", summary), ("Details", details)])
    return ReportResult(
        title=title,
        job_id=job_dir.name,
        xlsx_path=xlsx_path,
        html_path=html_path,
        pdf_path=pdf_path,
        added=len(added_keys),
        removed=len(removed_keys),
        duplication_add=new_duplicates,
        duplication_remove=old_duplicates,
        duplication=new_duplicates + old_duplicates,
        changed_rows=len(changed_keys),
        changed_fields=changed_fields,
    )


def generate_student_gantt_report(source: Path, output_root: Path) -> ReportResult:
    job_dir = _make_job_dir(output_root)
    frame = pd.read_excel(source)
    normalized = _normalize_gantt_columns(frame)
    normalized["start"] = pd.to_datetime(normalized["start"], errors="coerce")
    normalized["end"] = pd.to_datetime(normalized["end"], errors="coerce")
    normalized = normalized.dropna(subset=["student", "task", "start", "end"]).copy()
    normalized["days"] = (normalized["end"] - normalized["start"]).dt.days + 1
    normalized["progress"] = pd.to_numeric(normalized.get("progress", 0), errors="coerce").fillna(0).clip(0, 100)

    title = "Student Gantt Report"
    export = normalized[["student", "task", "start", "end", "days", "progress"]].rename(
        columns={
            "student": "Student",
            "task": "Task",
            "start": "Start",
            "end": "End",
            "days": "Days",
            "progress": "Progress",
        }
    )
    return _write_report(title, job_dir, [("Gantt", export)], _gantt_html(title, export))


def _make_job_dir(output_root: Path) -> Path:
    job_dir = output_root / uuid4().hex
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def _write_report(title: str, job_dir: Path, sheets: list[tuple[str, pd.DataFrame]], html: str) -> ReportResult:
    xlsx_path = job_dir / "report.xlsx"
    html_path = job_dir / "report.html"
    pdf_path = job_dir / "report.pdf"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sheet_name, frame in sheets:
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    html_path.write_text(html, encoding="utf-8")
    _write_pdf(pdf_path, title, sheets)
    return ReportResult(title=title, job_id=job_dir.name, xlsx_path=xlsx_path, html_path=html_path, pdf_path=pdf_path)


def _summary_html(title: str, source_label: str, summary: pd.DataFrame, details: pd.DataFrame | None = None) -> str:
    detail_html = "" if details is None or details.empty else f"<h2>Details</h2>{details.to_html(index=False, escape=True)}"
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>{escape(title)}</title>
  <style>{_css()}</style>
</head>
<body>
  <main>
    <h1>{escape(title)}</h1>
    <p class="muted">{escape(source_label)}</p>
    {summary.to_html(index=False, escape=True)}
    {detail_html}
  </main>
</body>
</html>"""


def _gantt_html(title: str, frame: pd.DataFrame) -> str:
    if frame.empty:
        rows = "<p>No valid gantt rows found.</p>"
    else:
        min_start = frame["Start"].min()
        max_end = frame["End"].max()
        total_days = max((max_end - min_start).days + 1, 1)
        rendered = []
        for _, row in frame.sort_values(["Student", "Start", "End"]).iterrows():
            offset = ((row["Start"] - min_start).days / total_days) * 100
            width = max((row["Days"] / total_days) * 100, 2)
            progress = int(row["Progress"])
            rendered.append(
                f"""<div class="gantt-row">
  <div class="label"><strong>{escape(str(row["Student"]))}</strong><span>{escape(str(row["Task"]))}</span></div>
  <div class="track"><div class="bar" style="left:{offset:.2f}%;width:{width:.2f}%"><span>{progress}%</span></div></div>
</div>"""
            )
        rows = "\n".join(rendered)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>{escape(title)}</title>
  <style>{_css()}</style>
</head>
<body>
  <main>
    <h1>{escape(title)}</h1>
    <section class="gantt">{rows}</section>
  </main>
</body>
</html>"""


def _write_pdf(path: Path, title: str, sheets: list[tuple[str, pd.DataFrame]]) -> None:
    doc = SimpleDocTemplate(str(path), pagesize=A4)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    for sheet_name, frame in sheets:
        story.append(Paragraph(sheet_name, styles["Heading2"]))
        preview = frame.head(20).fillna("").astype(str)
        data = [list(preview.columns)] + preview.values.tolist()
        table = Table(data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#172b4d")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dcdfe4")),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.extend([table, Spacer(1, 12)])
    doc.build(story)


def _matching_old_sheet(old_wb, new_sheet_name: str):
    if new_sheet_name in old_wb.sheetnames:
        return old_wb[new_sheet_name]
    if "Export" in old_wb.sheetnames:
        return old_wb["Export"]
    return old_wb.worksheets[-1]


def _headers(ws) -> list[str]:
    return ["" if cell.value is None else str(cell.value).strip() for cell in ws[HEADER_ROW]]


def _key_map(ws) -> tuple[dict[str, int], int]:
    rows_by_key: dict[str, int] = {}
    duplicates = 0
    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        key = _normalized_key(ws.cell(row=row_idx, column=KEY_COLUMN_INDEX).value)
        if not key:
            continue
        if key in rows_by_key:
            duplicates += 1
            continue
        rows_by_key[key] = row_idx
    return rows_by_key, duplicates


def _comparable_columns(old_headers: list[str], new_headers: list[str]) -> list[tuple[int, str]]:
    columns = []
    old_header_set = set(old_headers)
    for idx, header in enumerate(new_headers, start=1):
        if not header or header not in old_header_set:
            continue
        if header in EXCLUDED_COMPARE_COLUMNS:
            continue
        columns.append((idx, header))
    return columns


def _normalized_key(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalized_cell(value):
    if value is None:
        return ""
    if isinstance(value, str):
        stripped = value.strip()
        try:
            return float(stripped)
        except ValueError:
            return stripped
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def _write_compare_workbook(
    *,
    xlsx_path: Path,
    old_ws,
    new_ws,
    old_headers: list[str],
    new_headers: list[str],
    old_map: dict[str, int],
    new_map: dict[str, int],
    added_keys: set[str],
    removed_keys: set[str],
    changed_detail_rows: list[dict],
    summary: pd.DataFrame,
    details: pd.DataFrame,
) -> None:
    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    new_data_ws = workbook.create_sheet("new data")
    old_data_ws = workbook.create_sheet("old data")

    _write_summary_sheet(summary_ws, summary, details)
    _copy_values(new_ws, new_data_ws)
    _copy_values(old_ws, old_data_ws)

    new_header_set = set(new_headers)
    old_header_set = set(old_headers)
    excluded_new_columns = _column_indexes(new_headers, EXCLUDED_COMPARE_COLUMNS)
    excluded_old_columns = _column_indexes(old_headers, EXCLUDED_COMPARE_COLUMNS)

    for col_idx in excluded_new_columns:
        _fill_column(new_data_ws, col_idx, FILL_EXCLUDED)
    for col_idx in excluded_old_columns:
        _fill_column(old_data_ws, col_idx, FILL_EXCLUDED)
    for col_idx, header in enumerate(new_headers, start=1):
        if header and header not in old_header_set:
            new_data_ws.cell(row=HEADER_ROW, column=col_idx).fill = FILL_NEW_COLUMN

    for key in added_keys:
        _fill_row(new_data_ws, new_map[key], FILL_ADDED)
    for key in removed_keys:
        _fill_row(old_data_ws, old_map[key], FILL_REMOVED)
    for row in changed_detail_rows:
        if row["Type"] != "Changed":
            continue
        field = row["Field"]
        if field not in new_header_set:
            continue
        col_idx = new_headers.index(field) + 1
        new_data_ws.cell(row=row["Line No"], column=col_idx).fill = FILL_CHANGED

    for ws in (summary_ws, new_data_ws, old_data_ws):
        _auto_width(ws)
    workbook.save(xlsx_path)


def _write_summary_sheet(ws, summary: pd.DataFrame, details: pd.DataFrame) -> None:
    ws["A1"] = "Monthly Sales Plan Compare Summary"
    ws["A1"].font = Font(bold=True, size=14)
    for row_offset, row in enumerate(summary.itertuples(index=False), start=2):
        ws.cell(row=row_offset, column=1, value=row.Metric)
        ws.cell(row=row_offset, column=2, value=row.Count)
    detail_start = len(summary) + 4
    ws.cell(row=detail_start, column=1, value="Details")
    ws.cell(row=detail_start, column=1).font = Font(bold=True)
    for col_idx, column_name in enumerate(details.columns, start=1):
        cell = ws.cell(row=detail_start + 1, column=col_idx, value=column_name)
        cell.fill = FILL_HEADER
        cell.font = HEADER_FONT
    for row_idx, row in enumerate(details.itertuples(index=False), start=detail_start + 2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        fill = {"Added": FILL_ADDED, "Removed": FILL_REMOVED, "Changed": FILL_CHANGED}.get(row[0])
        if fill:
            for col_idx in range(1, len(details.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    legend_start = detail_start + max(len(details), 1) + 4
    legend = [("Yellow", "Added record"), ("Gray", "Removed record"), ("Orange", "Changed value"), ("Green", "New column")]
    ws.cell(row=legend_start, column=1, value="Color legend").font = Font(bold=True)
    fills = [FILL_ADDED, FILL_REMOVED, FILL_CHANGED, FILL_NEW_COLUMN]
    for idx, (color_name, meaning) in enumerate(legend, start=legend_start + 1):
        ws.cell(row=idx, column=1, value=color_name).fill = fills[idx - legend_start - 1]
        ws.cell(row=idx, column=2, value=meaning)


def _copy_values(source_ws, target_ws) -> None:
    for row in source_ws.iter_rows():
        for cell in row:
            target_ws.cell(row=cell.row, column=cell.column, value=cell.value)


def _column_indexes(headers: list[str], names: set[str]) -> list[int]:
    return [idx for idx, header in enumerate(headers, start=1) if header in names]


def _fill_row(ws, row_idx: int, fill: PatternFill) -> None:
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _fill_column(ws, col_idx: int, fill: PatternFill) -> None:
    for row_idx in range(1, ws.max_row + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 48)


def _normalize_gantt_columns(frame: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "student": ["student", "學生", "学生", "name", "student name", "氏名"],
        "task": ["task", "任務", "任务", "課題", "作業", "activity"],
        "start": ["start", "start date", "開始", "開始日", "開始日期"],
        "end": ["end", "end date", "結束", "終了", "結束日", "終了日"],
        "progress": ["progress", "進度", "进度", "完成率", "progress %"],
    }
    lowered = {str(column).strip().lower(): column for column in frame.columns}
    selected = {}
    for target, names in aliases.items():
        for name in names:
            if name.lower() in lowered:
                selected[target] = frame[lowered[name.lower()]]
                break
    missing = {"student", "task", "start", "end"} - set(selected)
    if missing:
        raise ValueError(f"Missing required gantt columns: {', '.join(sorted(missing))}")
    return pd.DataFrame(selected)


def _css() -> str:
    return """
body { margin: 0; background: #f7f8f9; color: #172b4d; font-family: Segoe UI, Arial, sans-serif; }
main { width: min(1120px, calc(100vw - 40px)); margin: 32px auto; }
h1 { font-size: 28px; margin: 0 0 8px; }
h2 { font-size: 18px; margin-top: 28px; }
.muted { color: #626f86; }
table { border-collapse: collapse; width: 100%; background: white; border: 1px solid #dcdfe4; }
th, td { border: 1px solid #dcdfe4; padding: 8px 10px; text-align: left; }
th { background: #172b4d; color: white; }
.gantt { display: grid; gap: 10px; }
.gantt-row { display: grid; grid-template-columns: 220px 1fr; gap: 12px; align-items: center; }
.label { background: white; border: 1px solid #dcdfe4; padding: 10px; border-radius: 6px; }
.label span { display: block; color: #626f86; margin-top: 3px; }
.track { position: relative; height: 34px; background: white; border: 1px solid #dcdfe4; border-radius: 6px; overflow: hidden; }
.bar { position: absolute; top: 5px; bottom: 5px; background: #0c66e4; border-radius: 4px; color: white; display: grid; place-items: center; font-size: 12px; }
"""
